#!/usr/bin/env python3
"""
Sprint 019 — Intacct Data: Start Fresh from DriveShop's Catalog.

Replaces the app's placeholder (test) items + per-client prices with DriveShop's
real 160-item catalog and Intacct coding, so the item foundation is correct and
the Intacct AR/AP export is unblocked at the source (every item ships with its
own Item ID + GL codes — nothing has to be matched).

Reads:
  data/imports/Item IDs - Dave M Edits_06.24.26.xlsx   (160-item catalog)
  data/imports/Intacct Coding.xlsx                     (offices, segments, customers)

Usage:
    python scripts/import_intacct_catalog.py            # dry-run, prints summary only
    python scripts/import_intacct_catalog.py --confirm  # also writes the .sql

Then an OPERATOR reviews and applies scripts/import_intacct_catalog.sql via the
Supabase SQL editor. This script NEVER writes to the database.

Approach (approved 2026-07-01) — FULL REPLACE, one atomic BEGIN/COMMIT, FK-safe order:
  A. Detach test-estimate soft-links (estimate_line_items / schedule_entries /
     labor_entries .rate_card_item_id -> NULL). Those rows already denormalize
     unit_rate / unit_cost / gl_code / name at creation, so detaching the FK does
     not change any stored estimate number.
  B. DELETE FROM rate_card_items  — clears test prices (rate cards stay EMPTY
     until DriveShop delivers real pricing in a future sprint).
  C. DELETE FROM fee_types        — removes the test item records (now unreferenced;
     fee_types has no is_active column, so a clean delete is the only fresh start).
  D. INSERT the 160 catalog items into fee_types with their Item ID + GL codes.
  E. Reference tables: office_accounting_profiles (15), revenue_segments (10),
     clients.intacct_customer_id on clean name matches.

Safety:
  - historical_events / historical_patterns store their data as JSON and have NO
    FK to fee_types / rate_card_items — they are untouched by this load.
  - No schema changes. No exporter edits. No invented prices. Office Payout from
    the catalog is NOT loaded (fee_types has no column for it; the app uses the
    per-client clients.office_payout_pct). Non-standard payout items are logged as
    a question for Dave, not written.

Column mapping (catalog -> fee_types):
  Item Name    -> name
  Section      -> section        (Planning & Admin/Onsite Labor/Travel/Creative/
                                   Production/Logistics -> planning_admin/onsite_labor/
                                   travel/creative/production/logistics)
  Cost Type    -> cost_type       (Labor/Flat Fee/Pass Through -> labor/flat_fee/pass_through)
  Revenue GL   -> gl_code
  Item ID      -> intacct_ar_item_id
  Cost GL      -> intacct_ap_gl_account_no
  GL Descript. -> accounting_memo
  (default_unit left to its DB default 'Each'; office payout NOT loaded)
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import openpyxl

# ---- Paths ----

PROJECT_ROOT = Path(__file__).resolve().parent.parent
CATALOG_PATH = PROJECT_ROOT / "data" / "imports" / "Item IDs - Dave M Edits_06.24.26.xlsx"
CODING_PATH = PROJECT_ROOT / "data" / "imports" / "Intacct Coding.xlsx"
OUTPUT_SQL_PATH = PROJECT_ROOT / "scripts" / "import_intacct_catalog.sql"

# ---- Mappings ----

SECTION_MAP = {
    "Planning & Admin": "planning_admin",
    "Onsite Labor": "onsite_labor",
    "Travel": "travel",
    "Creative": "creative",
    "Production": "production",
    "Logistics": "logistics",
}

COST_TYPE_MAP = {
    "Labor": "labor",
    "Flat Fee": "flat_fee",
    "Pass Through": "pass_through",
}

# Office payout is NOT loaded (fee_types has no column; the app uses the per-client
# clients.office_payout_pct). These are only used to log a question for Dave.
# Expected values: 0.75 = standard office split; 1.0 = pass-through / production /
# travel (offices don't take a cut of pass-throughs — see DOMAIN.md). Anything else
# (0.5, 0.9) is genuinely non-standard and worth confirming.
EXPECTED_OFFICE_PAYOUTS = {0.75, 1.0}

# Current app clients (DB snapshot 2026-07-01). Used only to decide which customer
# name matches are "clean"; the generated UPDATEs are still guarded by a name WHERE
# clause, so they no-op if a client has since been renamed/removed.
APP_CLIENTS = [
    "Acura", "Audi", "Bentley", "Ferrari", "Genesis", "Hankook", "Honda",
    "Hyundai", "JLR", "Lamborghini", "Lexus", "Lucid", "MB", "Maserati",
    "Mazda", "No Client", "Polestar", "Porsche", "Toyota", "VW", "Volkswagen",
    "Volvo", "Volvo MS",
]


def norm(s) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def gl_text(v) -> str | None:
    """Normalize a GL code cell to exact text.

    The Cost-GL column is numeric for 131/160 rows, so str(float) silently drops
    trailing zeros (5000.10 -> '5000.1'), which would produce a WRONG Intacct
    account number. Format floats to 2 decimals; 3-part sub-accounts (e.g.
    '5000.26.01') arrive as text and are passed through unchanged.
    """
    if v is None or v == "":
        return None
    if isinstance(v, float):
        return f"{v:.2f}"
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


# ---- SQL helpers (mirror import_rate_cards.py) ----


def sql_str(s) -> str:
    if s is None or s == "":
        return "NULL"
    return "'" + str(s).replace("'", "''") + "'"


# ---- Parsing ----


class CatalogItem:
    __slots__ = ("section", "name", "in_estimate", "item_id", "cost_type",
                 "rev_gl", "gl_desc", "cost_gl", "office_payout")

    def __init__(self, row):
        self.section = str(row[0]).strip() if row[0] is not None else None
        self.name = str(row[1]).strip() if row[1] is not None else None
        self.in_estimate = row[2]
        self.item_id = str(row[3]).strip() if row[3] is not None else None
        self.cost_type = str(row[4]).strip() if row[4] is not None else None
        self.rev_gl = gl_text(row[5])
        self.gl_desc = str(row[6]).strip() if row[6] is not None else None
        self.cost_gl = gl_text(row[8])
        self.office_payout = row[9]


def read_catalog() -> list[CatalogItem]:
    wb = openpyxl.load_workbook(CATALOG_PATH, data_only=True, read_only=True)
    try:
        ws = wb["Sheet1"]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    items = [CatalogItem(r) for r in rows[1:] if r[3] is not None]
    return items


def read_offices() -> list[dict]:
    wb = openpyxl.load_workbook(CODING_PATH, data_only=True, read_only=True)
    try:
        rows = list(wb["Vendors-Affiliates"].iter_rows(values_only=True))
    finally:
        wb.close()
    offices = []
    for r in rows[1:]:
        office = str(r[0]).strip() if r[0] is not None else None
        vendor_name = str(r[3]).strip() if r[3] is not None else None
        # Skip the blank Corporate row (no vendor) — not an AP vendor.
        if not office or vendor_name is None:
            continue
        offices.append({
            "office_name": office,
            "legal_name": vendor_name,
            "intacct_vendor_id": str(r[4]).strip() if r[4] is not None else None,
            "default_payment_terms": str(int(r[5])) if isinstance(r[5], (int, float)) else (str(r[5]).strip() if r[5] else None),
            "default_location_id": str(r[1]).strip() if r[1] is not None else None,
        })
    return offices


def read_segments() -> list[dict]:
    wb = openpyxl.load_workbook(CODING_PATH, data_only=True, read_only=True)
    try:
        rows = list(wb["Departments-Segment"].iter_rows(values_only=True))
    finally:
        wb.close()
    segs = []
    for r in rows[1:]:
        if r[0] is None and r[1] is None:
            continue
        code = str(int(r[0])) if isinstance(r[0], (int, float)) else (str(r[0]).strip() if r[0] else None)
        name = str(r[1]).strip() if r[1] is not None else None
        if not name:
            continue
        segs.append({"name": name, "code": code})
    return segs


def read_customers() -> dict[str, tuple[str, str]]:
    wb = openpyxl.load_workbook(CODING_PATH, data_only=True, read_only=True)
    try:
        rows = list(wb["Customers"].iter_rows(values_only=True))
    finally:
        wb.close()
    out = {}
    for r in rows[1:]:
        cid = str(r[0]).strip() if r[0] is not None else None
        cname = str(r[1]).strip() if r[1] is not None else None
        if cid and cname:
            out.setdefault(norm(cname), (cid, cname))
    return out


def match_clients(customers: dict[str, tuple[str, str]]):
    """Return (matches, exceptions). matches: list of (client_name, cust_id, cust_name)."""
    matches, exceptions = [], []
    for client in APP_CLIENTS:
        hit = customers.get(norm(client))
        if hit:
            matches.append((client, hit[0], hit[1]))
        else:
            exceptions.append(client)
    return matches, exceptions


# ---- SQL generation ----


def generate_sql(catalog, offices, segments, matches) -> str:
    L: list[str] = []
    L.append("-- =============================================================================")
    L.append("-- Sprint 019 — Intacct Data: Start Fresh from DriveShop's Catalog")
    L.append(f"-- Generated from: {CATALOG_PATH.name} + {CODING_PATH.name}")
    L.append("-- FULL REPLACE of test items/prices with the real 160-item catalog.")
    L.append("-- Atomic: the whole load is one transaction. Review before applying.")
    L.append("-- historical_events / historical_patterns are JSON-only and untouched.")
    L.append("-- =============================================================================")
    L.append("")
    L.append("BEGIN;")
    L.append("")

    # A. Detach test-estimate soft-links (denormalized values preserved).
    L.append("-- A. Detach rate_card_item_id soft-links on existing (test) estimates.")
    L.append("--    Estimate rows copy unit_rate/unit_cost/gl_code/name at creation, so")
    L.append("--    nulling the FK does not change any stored estimate number.")
    L.append("UPDATE estimate_line_items SET rate_card_item_id = NULL WHERE rate_card_item_id IS NOT NULL;")
    L.append("UPDATE schedule_entries    SET rate_card_item_id = NULL WHERE rate_card_item_id IS NOT NULL;")
    L.append("UPDATE labor_entries       SET rate_card_item_id = NULL WHERE rate_card_item_id IS NOT NULL;")
    L.append("")

    # B. Clear test prices.
    L.append("-- B. Clear the test per-client prices (rate cards stay EMPTY until real pricing).")
    L.append("DELETE FROM rate_card_items;")
    L.append("")

    # C. Remove test fee/item records.
    L.append("-- C. Remove the test fee/item records (now unreferenced).")
    L.append("DELETE FROM fee_types;")
    L.append("")

    # D. Load the 160 catalog items.
    L.append(f"-- D. Load the {len(catalog)} catalog items as the real item foundation.")
    L.append("INSERT INTO fee_types")
    L.append("  (name, section, cost_type, gl_code, intacct_ar_item_id, intacct_ap_gl_account_no, accounting_memo, display_order)")
    L.append("VALUES")
    value_rows = []
    for i, it in enumerate(catalog, start=1):
        section = SECTION_MAP[it.section]
        cost_type = COST_TYPE_MAP[it.cost_type]
        value_rows.append(
            f"  ({sql_str(it.name)}, {sql_str(section)}, {sql_str(cost_type)}, "
            f"{sql_str(it.rev_gl)}, {sql_str(it.item_id)}, {sql_str(it.cost_gl)}, "
            f"{sql_str(it.gl_desc)}, {i})"
        )
    L.append(",\n".join(value_rows) + ";")
    L.append("")

    # E. Reference tables.
    L.append(f"-- E1. office_accounting_profiles ({len(offices)}) — idempotent on office_name.")
    for o in offices:
        L.append(
            "INSERT INTO office_accounting_profiles "
            "(office_name, legal_name, intacct_vendor_id, default_payment_terms, default_location_id)"
        )
        L.append(
            f"  VALUES ({sql_str(o['office_name'])}, {sql_str(o['legal_name'])}, "
            f"{sql_str(o['intacct_vendor_id'])}, {sql_str(o['default_payment_terms'])}, "
            f"{sql_str(o['default_location_id'])})"
        )
        L.append(
            "  ON CONFLICT (office_name) DO UPDATE SET "
            "legal_name = EXCLUDED.legal_name, intacct_vendor_id = EXCLUDED.intacct_vendor_id, "
            "default_payment_terms = EXCLUDED.default_payment_terms, "
            "default_location_id = EXCLUDED.default_location_id, active = true;"
        )
    L.append("")

    L.append(f"-- E2. revenue_segments ({len(segments)}) — idempotent on name.")
    for i, s in enumerate(segments, start=1):
        L.append(
            "INSERT INTO revenue_segments (name, code, sort_order) "
            f"VALUES ({sql_str(s['name'])}, {sql_str(s['code'])}, {i}) "
            "ON CONFLICT (name) DO UPDATE SET code = EXCLUDED.code, sort_order = EXCLUDED.sort_order, active = true;"
        )
    L.append("")

    L.append(f"-- E3. clients.intacct_customer_id — {len(matches)} clean name matches (guarded).")
    for client, cid, cname in matches:
        L.append(
            f"UPDATE clients SET intacct_customer_id = {sql_str(cid)} "
            f"WHERE lower(name) = lower({sql_str(client)});  -- {cname}"
        )
    L.append("")

    L.append("COMMIT;")
    L.append("")
    return "\n".join(L)


# ---- Dry-run summary ----


def print_summary(catalog, offices, segments, matches, exceptions):
    print()
    print("=" * 74)
    print("Sprint 019 — Intacct Catalog Load — DRY RUN SUMMARY")
    print("=" * 74)

    # Catalog validation
    ids = [it.item_id for it in catalog]
    names = [norm(it.name) for it in catalog]
    dup_ids = [x for x in set(ids) if ids.count(x) > 1]
    dup_names = [x for x in set(names) if names.count(x) > 1]
    bad_section = [it.name for it in catalog if it.section not in SECTION_MAP]
    bad_cost = [it.name for it in catalog if it.cost_type not in COST_TYPE_MAP]
    missing = [it.item_id for it in catalog if not (it.item_id and it.rev_gl and it.cost_gl)]

    print(f"  Catalog items: {len(catalog)}")
    print(f"    duplicate Item IDs:   {dup_ids or 'NONE'}")
    print(f"    duplicate names:      {dup_names or 'NONE'}")
    print(f"    unmapped sections:    {bad_section or 'NONE'}")
    print(f"    unmapped cost types:  {bad_cost or 'NONE'}")
    print(f"    missing Item ID / Rev GL / Cost GL: {missing or 'NONE'}")

    if dup_ids or dup_names or bad_section or bad_cost or missing:
        print("  ⚠ Catalog validation FAILED — fix source before --confirm.")

    # Office payout: NOT loaded (no column). Only the genuinely non-standard values
    # (not 0.75 standard, not 1.0 pass-through) are logged as a question for Dave.
    from collections import Counter
    payout_dist = dict(Counter(it.office_payout for it in catalog))
    outliers = [it for it in catalog if it.office_payout not in EXPECTED_OFFICE_PAYOUTS]
    print(f"\n  Office payout (NOT loaded — no fee_types column). Distribution: {payout_dist}")
    print("    0.75 = standard office split; 1.0 = pass-through/production/travel (expected).")
    print(f"    NON-STANDARD items to confirm with Dave: {len(outliers)}")
    for it in sorted(outliers, key=lambda x: x.office_payout):
        print(f"    payout={it.office_payout}  {it.item_id}  {it.name}  [{it.section}]")

    print("\n  --- Writes the generated SQL will make (inside one transaction) ---")
    print("  A. detach rate_card_item_id on estimate_line_items / schedule_entries / labor_entries")
    print("     (full-table nulling of non-null soft-links on the existing test estimates)")
    print("  B. DELETE FROM rate_card_items   (all test prices)")
    print("  C. DELETE FROM fee_types         (all test item records)")
    print(f"  D. INSERT {len(catalog)} catalog items into fee_types (each carries intacct_ar_item_id)")
    print(f"  E1. upsert {len(offices)} office_accounting_profiles")
    print(f"  E2. upsert {len(segments)} revenue_segments")
    print(f"  E3. set intacct_customer_id on {len(matches)} clients (clean matches)")

    print(f"\n  Client → customer CLEAN matches ({len(matches)}):")
    for client, cid, cname in matches:
        print(f"    {client:<12} -> {cid}  ({cname})")

    print(f"\n  Client → customer EXCEPTIONS for the Dave meeting ({len(exceptions)}):")
    print(f"    {', '.join(exceptions)}")
    print("=" * 74)


def main() -> int:
    parser = argparse.ArgumentParser(description="Load DriveShop's Intacct catalog + coding (fresh start).")
    parser.add_argument("--confirm", action="store_true",
                        help="Write scripts/import_intacct_catalog.sql in addition to the dry-run summary.")
    args = parser.parse_args()

    for p in (CATALOG_PATH, CODING_PATH):
        if not p.exists():
            print(f"ERROR: source not found at {p}", file=sys.stderr)
            return 2

    catalog = read_catalog()
    offices = read_offices()
    segments = read_segments()
    customers = read_customers()
    matches, exceptions = match_clients(customers)

    print_summary(catalog, offices, segments, matches, exceptions)

    # Hard stop if the catalog does not validate — never emit a bad load.
    ids = [it.item_id for it in catalog]
    names = [norm(it.name) for it in catalog]
    ok = (len(catalog) == 160
          and len(set(ids)) == len(ids)
          and len(set(names)) == len(names)
          and all(it.section in SECTION_MAP and it.cost_type in COST_TYPE_MAP for it in catalog)
          and all(it.item_id and it.rev_gl and it.cost_gl for it in catalog))
    if not ok:
        print("\nABORT: catalog validation failed; not writing SQL.", file=sys.stderr)
        return 1

    if not args.confirm:
        print()
        print(f"Dry-run only. Re-run with --confirm to write {OUTPUT_SQL_PATH.relative_to(PROJECT_ROOT)}.")
        return 0

    sql = generate_sql(catalog, offices, segments, matches)
    OUTPUT_SQL_PATH.write_text(sql)
    print()
    print(f"✓ Wrote {OUTPUT_SQL_PATH.relative_to(PROJECT_ROOT)} ({len(sql.splitlines())} lines).")
    print("  Review, then apply via the Supabase SQL editor. This script does NOT apply it.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
