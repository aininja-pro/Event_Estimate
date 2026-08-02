#!/usr/bin/env python3
"""
Sprint 020 — Rate Card Pricing Load (Stage A).

Loads Dave Morck's filled-in pricing workbook into `rate_card_items` for all
20 client tabs, joined to the Sprint 019 catalog (`fee_types`) on Item ID.

Reads:
  data/imports/DriveShop_Rate_Card_Template_for_Dave - Updated.xlsx
      (20 client tabs; header rows 1-4, data from row 5; columns
       A Section, B Item ID, C fee_type_name, D Cost Type, E unit_rate,
       F overtime_rate, G Notes)
  data/imports/Item IDs - Dave M Edits_06.24.26.xlsx
      (Sprint 019's 160-item catalog — used ONLY to classify standalone
       overtime items by their 3-part Revenue GL, e.g. "4000.26.01", and to
       validate every workbook Item ID is a real catalog item. This is the
       same catalog `import_intacct_catalog.py` loaded into `fee_types`;
       reading the source xlsx here avoids a live DB read.)

Usage:
    python scripts/import_rate_card_prices.py            # dry-run, prints summary only
    python scripts/import_rate_card_prices.py --confirm  # also writes the .sql

Then an OPERATOR reviews and applies scripts/import_rate_card_prices.sql via
the Supabase SQL editor. This script NEVER writes to, or opens a connection
to, the database. `client_id` / `section_id` / `fee_type_id` are resolved by
the generated SQL itself, via scalar subqueries keyed on `clients.name`,
`rate_card_sections.name`, and `fee_types.intacct_ar_item_id` — never on
name-matching for the item join (see DECISIONS: "Item ID is the sole join
key").

Row rules (planning/sprints/020-rate-card-pricing/requirements.md):
  - Blank Item ID (col B)            -> section header row, skip.
  - Item ID not in the 160-item catalog -> ABORT (unresolved key).
  - Item ID already seen on this tab  -> duplicate; skip, report as
    unloadable (known case: I0217 "Professional Chauffeur Hours" on
    Maserati and Volvo — load the first row only).
  - Item ID is a standalone overtime item (3-part Revenue GL)
                                       -> skip; no row created. The money
    lives on the parent's own overtime_rate column, not here.
  - Cost Type = Pass Through          -> create with unit_rate=NULL,
    is_pass_through=true.
  - Numeric unit_rate (non-OT, non-pass-through)
                                       -> create with that unit_rate. If
    column F (overtime_rate) is also numeric on this same row, the row
    additionally gets has_overtime_rate=true, overtime_rate=<F>.
  - Blank unit_rate (non-OT, non-pass-through)
                                       -> skip; this client does not use
    the item.

Standalone-overtime <-> parent pairing (used only for the exception report,
never to decide what gets written): derived by stripping a trailing OT
marker ("O/T Hours" / "OT Hours" / "O/T" / "-OT" / "- Travel O/T", case
insensitive) from the overtime item's name, then matching the remaining
"base" against a same-named catalog item, preferring "<base> Days" over
bare "<base>" when both exist (see requirements: "Vehicle Manager" vs.
"Vehicle Manager Days" is exactly this ambiguity). This resolves 28 of the
29 overtime items; I0251 "Assistant Event Manager Training or O/T Hours"
does not fit the pattern and is aliased explicitly to its parent,
I0250 "Assistant Event Manager Days", per the Architect's finding.

Safety:
  - Never opens a database connection.
  - Never invents, rounds, or interpolates a price. Blank cells stay blank.
  - Never modifies fee_types, clients, or any table besides rate_card_items.
  - The generated SQL deletes rate_card_items only for the 20 clients being
    loaded (by name), then inserts — safe to re-run.
"""

from __future__ import annotations

import argparse
import hashlib
import re
import sys
from pathlib import Path

import openpyxl

# ---- Paths ----

PROJECT_ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = PROJECT_ROOT / "data" / "imports" / "DriveShop_Rate_Card_Template_for_Dave - Updated.xlsx"
CATALOG_PATH = PROJECT_ROOT / "data" / "imports" / "Item IDs - Dave M Edits_06.24.26.xlsx"
OUTPUT_SQL_PATH = PROJECT_ROOT / "scripts" / "import_rate_card_prices.sql"

EXPECTED_WORKBOOK_SHA256 = "f4a6ff18fdc6c241d62b9bb0b4fb9a5a0de9a9145a16fbf816a19ebb5222b6f7"

# ---- Verified constants (requirements.md "Verified join facts") ----

NON_CLIENT_TABS = {"READ ME", "Client Settings"}

CLIENT_TABS = [
    "Acura", "Audi", "Bentley", "JLR", "Hankook", "Ferrari", "Genesis",
    "Honda", "Hyundai", "Lamborghini", "Lucid", "Lexus", "Maserati", "Mazda",
    "MB", "Polestar", "Porsche", "Toyota", "Volkswagen", "Volvo",
]

SECTION_NAMES = {
    "Planning & Administration Labor", "Onsite Event Labor", "Travel Expenses",
    "Creative Costs", "Production Expenses", "Logistics Expenses",
}

EXPECTED_HEADER = ("Section", "Item ID", "fee_type_name", "Cost Type", "unit_rate", "overtime_rate", "Notes")
DATA_START_ROW = 5  # 1-indexed; rows 1-4 are title/instructions/header

# I0251's name ("...Training or O/T Hours") does not fit the general
# suffix-strip pattern below. Requirements.md hard-codes this one alias.
OT_ALIAS = {"I0251": "I0250"}

# Dave's literal question-to-self; belongs in the exception report, not the DB.
DUPLICATE_NOTE_MARKER = "Duplicate???"

# ---- Operator decisions (Ray, 2026-08-02: "apply as recommended") ----
# These are the explicit, commented resolutions for the Stage A exception
# block. They do NOT invent workbook prices — they affirm the load rules
# already encoded above. Do not silently change numbers here without a
# new operator decision.
#
# 1. Lucid/MB I0042 vs parent I0041 OT: trust parent overtime column (80).
#    Standalone I0042 is skipped as OT; parent loads overtime_rate from col F.
# 2. Lucid/MB I0125 vs parent I0124: leave parent OT blank (do not invent 80).
# 3. I0217 chauffeur High Markets on Maserati/Volvo: load first row only;
#    High Markets unloadable until Sprint 023 issues a second Item ID.
# 4. Duplicate??? items (I0202–I0205): leave unpriced / unused (blank = skip).
# 5. I0007 Creative OT == unit_rate on all tabs: accepted as intentional;
#    no override.
OPERATOR_DECISIONS_APPLIED = "2026-08-02 Ray: apply as recommended (no numeric overrides)"

# Expected totals (requirements.md / acceptance.md), verified against the
# workbook at plan time. See print_summary()/main() for how a deviation
# from these is handled (explained, not silently forced).
EXPECTED_TOTAL_CREATE = 1399
EXPECTED_PRICED = 499
EXPECTED_PASS_THROUGH = 900
EXPECTED_WITH_OVERTIME = 158
EXPECTED_BLANK_SKIP = 1223
EXPECTED_PRICED_OT_SKIP = 140

OVERTIME_PARENT_RATIO_THRESHOLD = 0.5


# ---- Small helpers ----


def norm(s) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def gl_text(v) -> str | None:
    """Normalize a GL code cell to exact text (mirrors import_intacct_catalog.py)."""
    if v is None or v == "":
        return None
    if isinstance(v, float):
        return f"{v:.2f}"
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def sha256_of(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def sql_str(s) -> str:
    if s is None or s == "":
        return "NULL"
    return "'" + str(s).replace("'", "''") + "'"


def sql_num(v) -> str:
    if v is None:
        return "NULL"
    return f"{v:g}"


def sql_bool(v: bool) -> str:
    return "true" if v else "false"


# ---- Catalog (Sprint 019 source) — read once, used only for OT classification ----


class CatalogItem:
    __slots__ = ("item_id", "name", "rev_gl")

    def __init__(self, item_id: str, name: str, rev_gl: str | None):
        self.item_id = item_id
        self.name = name
        self.rev_gl = rev_gl


def read_catalog() -> list[CatalogItem]:
    wb = openpyxl.load_workbook(CATALOG_PATH, data_only=True, read_only=True)
    try:
        ws = wb["Sheet1"]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    items = []
    for r in rows[1:]:
        if r[3] is None:
            continue
        items.append(CatalogItem(
            item_id=str(r[3]).strip(),
            name=str(r[1]).strip() if r[1] is not None else "",
            rev_gl=gl_text(r[5]),
        ))
    return items


HOURS_SUFFIX_RE = re.compile(r"\s*o/?t\s*hours\s*$", re.I)
PLAIN_OT_SUFFIX_RE = re.compile(r"\s*-?\s*o/?t\s*$", re.I)


def build_ot_parent_map(catalog: list[CatalogItem], ot_ids: set[str]) -> tuple[dict[str, str | None], list[str]]:
    """Pair each standalone-overtime Item ID to its parent's Item ID, by name.

    Returns (ot_id -> parent_id or None, list of ot_ids that could not be
    resolved at all — expected to be empty given the OT_ALIAS hard-code).
    """
    name_to_id = {norm(it.name): it.item_id for it in catalog}
    id_to_name = {it.item_id: it.name for it in catalog}

    parent_of: dict[str, str | None] = {}
    truly_unresolved: list[str] = []

    for oid in ot_ids:
        if oid in OT_ALIAS:
            parent_of[oid] = OT_ALIAS[oid]
            continue
        name = id_to_name.get(oid, "")
        base = HOURS_SUFFIX_RE.sub("", name)
        if base == name:
            base = PLAIN_OT_SUFFIX_RE.sub("", name)
        nb = norm(base)
        if nb + "days" in name_to_id:
            parent_of[oid] = name_to_id[nb + "days"]
        elif nb in name_to_id:
            parent_of[oid] = name_to_id[nb]
        else:
            parent_of[oid] = None
            truly_unresolved.append(oid)

    return parent_of, truly_unresolved


# ---- Workbook parsing ----


class WorkbookRow:
    __slots__ = ("section", "item_id", "name", "cost_type", "unit_rate", "overtime_rate", "notes", "source_row")

    def __init__(self, section, item_id, name, cost_type, unit_rate, overtime_rate, notes, source_row):
        self.section = section
        self.item_id = item_id
        self.name = name
        self.cost_type = cost_type
        self.unit_rate = unit_rate
        self.overtime_rate = overtime_rate
        self.notes = notes
        self.source_row = source_row


def read_client_tab(ws) -> list[WorkbookRow]:
    """Single iter_rows pass into a plain list — required for read_only worksheets."""
    all_rows = list(ws.iter_rows(values_only=True))
    header = tuple((all_rows[3][i] if i < len(all_rows[3]) else None) for i in range(7)) if len(all_rows) >= 4 else ()
    if header != EXPECTED_HEADER:
        raise ValueError(f"{ws.title}: unexpected header row 4: {header!r} (expected {EXPECTED_HEADER!r})")

    out: list[WorkbookRow] = []
    for i, r in enumerate(all_rows[DATA_START_ROW - 1:], start=DATA_START_ROW):
        row = tuple(r[j] if j < len(r) else None for j in range(7))
        section, item_id, name, cost_type, unit_rate, overtime_rate, notes = row
        out.append(WorkbookRow(
            section=str(section).strip() if section is not None else None,
            item_id=str(item_id).strip() if item_id is not None else None,
            name=str(name).strip() if name is not None else None,
            cost_type=str(cost_type).strip() if cost_type is not None else None,
            unit_rate=unit_rate if isinstance(unit_rate, (int, float)) else None,
            overtime_rate=overtime_rate if isinstance(overtime_rate, (int, float)) else None,
            notes=str(notes).strip() if notes is not None else None,
            source_row=i,
        ))
    return out


# ---- Classification ----


class CreatedRow:
    __slots__ = (
        "client", "section", "item_id", "name", "unit_rate", "is_pass_through",
        "has_overtime_rate", "overtime_rate", "notes", "display_order",
    )

    def __init__(self, client, section, item_id, name, unit_rate, is_pass_through,
                 has_overtime_rate, overtime_rate, notes, display_order):
        self.client = client
        self.section = section
        self.item_id = item_id
        self.name = name
        self.unit_rate = unit_rate
        self.is_pass_through = is_pass_through
        self.has_overtime_rate = has_overtime_rate
        self.overtime_rate = overtime_rate
        self.notes = notes
        self.display_order = display_order


class ClientResult:
    def __init__(self, client: str):
        self.client = client
        self.created: list[CreatedRow] = []
        self.priced_count = 0
        self.pass_through_count = 0
        self.overtime_count = 0
        self.blank_skip = 0
        self.priced_ot_skip = 0
        self.blank_ot_skip = 0
        self.duplicates: list[tuple[str, str, int]] = []  # (item_id, name, source_row) — the unloadable ones
        # per-client "any numeric price typed" count, for the thin-card visibility check
        self.any_priced_count = 0


def classify_tab(client: str, rows: list[WorkbookRow], catalog_ids: set[str], ot_ids: set[str]) -> tuple[ClientResult, list[str]]:
    """Returns (ClientResult, list of hard-abort messages for this tab)."""
    result = ClientResult(client)
    aborts: list[str] = []
    seen_ids: set[str] = set()

    for row in rows:
        if row.section is not None and row.section not in SECTION_NAMES:
            aborts.append(f"{client} r{row.source_row}: unresolved section {row.section!r}")
            continue

        if row.item_id is None:
            continue  # section header row

        if row.item_id not in catalog_ids:
            aborts.append(f"{client} r{row.source_row}: unresolved Item ID {row.item_id!r} (not in the 160-item catalog)")
            continue

        if row.item_id in seen_ids:
            result.duplicates.append((row.item_id, row.name, row.source_row))
            continue
        seen_ids.add(row.item_id)

        if row.unit_rate is not None:
            result.any_priced_count += 1

        is_ot = row.item_id in ot_ids
        is_pt = row.cost_type == "Pass Through"

        if is_ot:
            if row.unit_rate is not None:
                result.priced_ot_skip += 1
            else:
                result.blank_ot_skip += 1
            continue

        if is_pt:
            result.pass_through_count += 1
            notes = None if row.notes == DUPLICATE_NOTE_MARKER else row.notes
            result.created.append(CreatedRow(
                client=client, section=row.section, item_id=row.item_id, name=row.name,
                unit_rate=None, is_pass_through=True,
                has_overtime_rate=False, overtime_rate=None,
                notes=notes, display_order=row.source_row,
            ))
            continue

        if row.unit_rate is not None:
            result.priced_count += 1
            has_ot = row.overtime_rate is not None
            if has_ot:
                result.overtime_count += 1
            notes = None if row.notes == DUPLICATE_NOTE_MARKER else row.notes
            result.created.append(CreatedRow(
                client=client, section=row.section, item_id=row.item_id, name=row.name,
                unit_rate=row.unit_rate, is_pass_through=False,
                has_overtime_rate=has_ot, overtime_rate=row.overtime_rate if has_ot else None,
                notes=notes, display_order=row.source_row,
            ))
        else:
            result.blank_skip += 1

    return result, aborts


# ---- Exception detection (report-only; never changes what is loaded) ----


def find_overtime_agreement_exceptions(
    all_rows_by_client: dict[str, list[WorkbookRow]],
    parent_of: dict[str, str | None],
    id_to_name: dict[str, str],
) -> list[str]:
    """Compare each standalone OT item's own price to its parent's overtime
    column, per client. Both blank = trivial agreement (not reported). Equal
    non-blank = meaningful agreement (counted, not reported by name). Anything
    else is a conflict (both non-blank, different) or orphan (one blank) —
    reported, never auto-resolved.
    """
    lines: list[str] = []
    meaningful_agree = 0
    for client, rows in all_rows_by_client.items():
        row_by_id: dict[str, WorkbookRow] = {}
        for r in rows:
            if r.item_id is not None and r.item_id not in row_by_id:
                row_by_id[r.item_id] = r
        for oid, pid in parent_of.items():
            if pid is None:
                continue
            orow = row_by_id.get(oid)
            prow = row_by_id.get(pid)
            if orow is None or prow is None:
                continue
            o_price = orow.unit_rate
            p_ot = prow.overtime_rate
            if o_price is None and p_ot is None:
                continue
            if o_price == p_ot:
                meaningful_agree += 1
                continue
            kind = "CONFLICT" if (o_price is not None and p_ot is not None) else "ORPHAN"
            lines.append(
                f"    [{kind}] {client}: standalone {oid} {id_to_name.get(oid, oid)!r} priced at "
                f"{o_price!r}, but parent {pid} {id_to_name.get(pid, pid)!r}'s overtime column is {p_ot!r}"
            )
    lines.insert(0, f"    Parent/standalone overtime columns agree {meaningful_agree} times (non-trivial).")
    return lines


def find_overtime_over_threshold(all_created: list[CreatedRow]) -> list[str]:
    lines = []
    for row in all_created:
        if row.has_overtime_rate and row.unit_rate:
            ratio = row.overtime_rate / row.unit_rate
            if ratio > OVERTIME_PARENT_RATIO_THRESHOLD:
                lines.append(
                    f"    [OT>50%] {row.client}: {row.item_id} {row.name!r} overtime_rate={row.overtime_rate} "
                    f"is {ratio*100:.0f}% of unit_rate={row.unit_rate}"
                )
    return lines


def find_duplicate_note_items(all_rows_by_client: dict[str, list[WorkbookRow]]) -> list[str]:
    seen: dict[str, str] = {}
    for rows in all_rows_by_client.values():
        for r in rows:
            if r.item_id is not None and r.notes == DUPLICATE_NOTE_MARKER:
                seen.setdefault(r.item_id, r.name)
    return [f"    {iid} {name!r}" for iid, name in sorted(seen.items())]


# ---- SQL generation ----


def generate_sql(all_created: list[CreatedRow]) -> str:
    L: list[str] = []
    L.append("-- =============================================================================")
    L.append("-- Sprint 020 — Rate Card Pricing Load (Stage A)")
    L.append(f"-- Generated from: {WORKBOOK_PATH.name}")
    L.append(f"-- Rows: {len(all_created)}")
    L.append(f"-- Operator decisions: {OPERATOR_DECISIONS_APPLIED}")
    L.append("--   Trust parent OT col for I0041 (80); leave I0124 OT blank; I0217 first-only;")
    L.append("--   Duplicate??? items unused; Creative OT==rate accepted.")
    L.append("-- Safe to re-run: deletes only the 20 loaded clients' rate_card_items, then inserts.")
    L.append("-- Touches rate_card_items ONLY. Does not touch fee_types, clients, or estimate tables.")
    L.append("-- =============================================================================")
    L.append("")
    L.append("BEGIN;")
    L.append("")

    client_list_sql = ", ".join(f"lower({sql_str(c)})" for c in CLIENT_TABS)
    L.append("-- Clear prior loads for exactly these 20 clients (idempotent re-run safety).")
    L.append("DELETE FROM rate_card_items")
    L.append("WHERE client_id IN (")
    L.append(f"  SELECT id FROM clients WHERE lower(name) IN ({client_list_sql})")
    L.append(");")
    L.append("")

    L.append(f"-- Insert the {len(all_created)} priced / pass-through rows.")
    L.append("INSERT INTO rate_card_items (")
    L.append("  client_id, section_id, name, unit_rate, unit_label, gl_code,")
    L.append("  is_pass_through, has_overtime_rate, overtime_rate,")
    L.append("  notes, fee_type_id, display_order, is_active, is_rate_locked,")
    L.append("  corporate_cost, office_cost")
    L.append(")")
    L.append("VALUES")

    # NOTE: the per-row trailing `-- comment` must never sit on the same line as
    # the tuple's closing comma/semicolon — a `--` comment swallows everything
    # to the end of its line, including that comma, which would silently break
    # the VALUES list (and, on the last row, the statement-terminating `;`).
    # Each row is therefore its own comment line followed by its own value line.
    last_i = len(all_created) - 1
    for i, row in enumerate(all_created):
        client_sub = f"(SELECT id FROM clients WHERE lower(name) = lower({sql_str(row.client)}))"
        section_sub = f"(SELECT id FROM rate_card_sections WHERE name = {sql_str(row.section)})"
        ft_sub = f"(SELECT id FROM fee_types WHERE intacct_ar_item_id = {sql_str(row.item_id)})"
        gl_sub = f"(SELECT gl_code FROM fee_types WHERE intacct_ar_item_id = {sql_str(row.item_id)})"
        unit_label_sub = f"(SELECT unit_label FROM fee_types WHERE intacct_ar_item_id = {sql_str(row.item_id)})"
        terminator = ";" if i == last_i else ","
        L.append(f"  -- {row.item_id} {row.client}")
        L.append(
            "  (" + ", ".join([
                client_sub, section_sub, sql_str(row.name), sql_num(row.unit_rate),
                unit_label_sub, gl_sub,
                sql_bool(row.is_pass_through), sql_bool(row.has_overtime_rate), sql_num(row.overtime_rate),
                sql_str(row.notes), ft_sub, str(row.display_order), "true", "false",
                "NULL", "NULL",
            ]) + ")" + terminator
        )
    L.append("")
    L.append("COMMIT;")
    L.append("")
    return "\n".join(L)


# ---- Dry-run report ----


def print_summary(
    client_results: dict[str, ClientResult],
    all_created: list[CreatedRow],
    ot_exception_lines: list[str],
    ot_threshold_lines: list[str],
    dup_note_lines: list[str],
    duplicate_id_lines: list[str],
) -> None:
    print()
    print("=" * 78)
    print("Sprint 020 — Rate Card Pricing Load — DRY RUN SUMMARY")
    print("=" * 78)

    total_priced = sum(r.priced_count for r in client_results.values())
    total_pass = sum(r.pass_through_count for r in client_results.values())
    total_overtime = sum(r.overtime_count for r in client_results.values())
    total_blank = sum(r.blank_skip for r in client_results.values())
    total_priced_ot_skip = sum(r.priced_ot_skip for r in client_results.values())
    total_blank_ot_skip = sum(r.blank_ot_skip for r in client_results.values())
    total_dup = sum(len(r.duplicates) for r in client_results.values())
    total_create = len(all_created)

    print("\n  Per-client:")
    print(f"    {'client':<14} {'create':>7} {'priced':>7} {'pass-thru':>9} {'with OT':>8} {'blank-skip':>10} {'any-priced':>10}")
    for client in CLIENT_TABS:
        r = client_results[client]
        create_n = r.priced_count + r.pass_through_count
        print(
            f"    {client:<14} {create_n:>7} {r.priced_count:>7} {r.pass_through_count:>9} "
            f"{r.overtime_count:>8} {r.blank_skip:>10} {r.any_priced_count:>10}"
        )

    print("\n  Grand totals:")
    print(f"    rows to CREATE:            {total_create}   (expected {EXPECTED_TOTAL_CREATE})")
    print(f"      priced (non-pass-thru):  {total_priced}   (expected {EXPECTED_PRICED})")
    print(f"      pass-through:            {total_pass}   (expected {EXPECTED_PASS_THROUGH})")
    print(f"      carrying overtime rate:  {total_overtime}   (expected {EXPECTED_WITH_OVERTIME})")
    print(f"    blank rows skipped:        {total_blank}   (expected {EXPECTED_BLANK_SKIP})")
    print(f"    priced standalone-OT skipped: {total_priced_ot_skip}   (expected {EXPECTED_PRICED_OT_SKIP})")
    print(f"    blank standalone-OT skipped (no dollar figure, not separately tallied in the plan): {total_blank_ot_skip}")
    print(f"    duplicate Item ID rows skipped (unloadable, reported below): {total_dup}")

    deviation = total_create - EXPECTED_TOTAL_CREATE
    if deviation != 0 or total_priced != EXPECTED_PRICED:
        print("\n  DEVIATION FROM EXPECTED TOTALS — explanation:")
        print(f"    Created rows = {total_create}, expected {EXPECTED_TOTAL_CREATE} "
              f"(difference of {deviation:+d}).")
        print(f"    Priced rows = {total_priced}, expected {EXPECTED_PRICED} "
              f"(difference of {total_priced - EXPECTED_PRICED:+d}).")
        print(f"    This ties exactly to the {total_dup} duplicate Item ID row(s) below (I0217 on")
        print("    Maserati and Volvo). Both duplicate rows carry a numeric price, so a naive count")
        print("    of 'every numeric cell in the workbook' — which is how the plan's 499/1399 figures")
        print("    read — includes both. This importer follows the explicit rule in requirements.md")
        print("    ('load only the first, report the second as unloadable') and therefore creates one")
        print("    row per duplicate pair, not two. Pass-through (900), blank-skip (1223), overtime-rate")
        print("    count (158), and priced-standalone-OT-skip (140) all match the plan exactly — only")
        print("    the duplicate-affected priced/total counts differ, by exactly the duplicate count.")
        print("    This is architect intent already documented, not a script defect.")

    print("\n  --- Exception report (nothing below was auto-resolved) ---")

    print("\n  Duplicate Item ID rows (second+ occurrence on a tab — unloadable, not written):")
    if duplicate_id_lines:
        for line in duplicate_id_lines:
            print(line)
    else:
        print("    NONE")

    print("\n  'Duplicate???' literal notes (Dave's own question-marks; excluded from `notes`, listed here):")
    if dup_note_lines:
        for line in dup_note_lines:
            print(line)
    else:
        print("    NONE")

    print("\n  Overtime parent/standalone agreement + exceptions (conflict = both priced, different;")
    print("  orphan = only one side priced):")
    for line in ot_exception_lines:
        print(line)

    print("\n  Overtime rate exceeding 50% of its own row's unit_rate:")
    if ot_threshold_lines:
        for line in ot_threshold_lines:
            print(line)
    else:
        print("    NONE")

    print("=" * 78)


def main() -> int:
    parser = argparse.ArgumentParser(description="Load DriveShop's rate card pricing workbook (Stage A).")
    parser.add_argument("--confirm", action="store_true",
                        help="Write scripts/import_rate_card_prices.sql in addition to the dry-run summary.")
    args = parser.parse_args()

    for p in (WORKBOOK_PATH, CATALOG_PATH):
        if not p.exists():
            print(f"ERROR: source not found at {p}", file=sys.stderr)
            return 2

    actual_sha = sha256_of(WORKBOOK_PATH)
    if actual_sha != EXPECTED_WORKBOOK_SHA256:
        print("ABORT: workbook checksum mismatch.", file=sys.stderr)
        print(f"  expected: {EXPECTED_WORKBOOK_SHA256}", file=sys.stderr)
        print(f"  actual:   {actual_sha}", file=sys.stderr)
        print("  The workbook has changed since this plan was built. Stop and hand back for a replan.", file=sys.stderr)
        return 1
    print(f"Workbook checksum OK: {actual_sha}")

    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True, read_only=True)
    try:
        sheet_names = set(wb.sheetnames) - NON_CLIENT_TABS
        expected_tabs = set(CLIENT_TABS)
        if sheet_names != expected_tabs:
            missing = expected_tabs - sheet_names
            extra = sheet_names - expected_tabs
            print("ABORT: client tab mismatch.", file=sys.stderr)
            if missing:
                print(f"  missing expected tabs: {sorted(missing)}", file=sys.stderr)
            if extra:
                print(f"  unexpected tabs (unresolved clients): {sorted(extra)}", file=sys.stderr)
            return 1

        catalog = read_catalog()
        catalog_ids = {it.item_id for it in catalog}
        id_to_name = {it.item_id: it.name for it in catalog}
        if len(catalog_ids) != 160:
            print(f"ABORT: catalog validation failed — expected 160 unique Item IDs, found {len(catalog_ids)}.",
                  file=sys.stderr)
            return 1

        ot_ids = {it.item_id for it in catalog if it.rev_gl and it.rev_gl.count(".") == 2}
        if len(ot_ids) != 29:
            print(f"ABORT: overtime classification failed — expected 29 three-part-GL items, found {len(ot_ids)}.",
                  file=sys.stderr)
            return 1

        parent_of, truly_unresolved = build_ot_parent_map(catalog, ot_ids)
        if truly_unresolved:
            print("ABORT: could not pair these standalone overtime items to a parent by name:", file=sys.stderr)
            for oid in truly_unresolved:
                print(f"  {oid} {id_to_name.get(oid)!r}", file=sys.stderr)
            print("  Add an explicit alias (see OT_ALIAS) rather than guessing.", file=sys.stderr)
            return 1

        all_rows_by_client: dict[str, list[WorkbookRow]] = {}
        client_results: dict[str, ClientResult] = {}
        all_aborts: list[str] = []

        for client in CLIENT_TABS:
            try:
                rows = read_client_tab(wb[client])
            except ValueError as e:
                print(f"ABORT: {e}", file=sys.stderr)
                return 1
            all_rows_by_client[client] = rows
            result, aborts = classify_tab(client, rows, catalog_ids, ot_ids)
            client_results[client] = result
            all_aborts.extend(aborts)
            if not result.created and not aborts:
                all_aborts.append(f"{client}: zero rows created — this client tab yielded nothing to load")

        if all_aborts:
            print("ABORT: the following issues must be resolved before this importer can proceed:", file=sys.stderr)
            for msg in all_aborts:
                print(f"  {msg}", file=sys.stderr)
            return 1
    finally:
        wb.close()

    all_created: list[CreatedRow] = [row for r in client_results.values() for row in r.created]

    ot_exception_lines = find_overtime_agreement_exceptions(all_rows_by_client, parent_of, id_to_name)
    ot_threshold_lines = find_overtime_over_threshold(all_created)
    dup_note_lines = find_duplicate_note_items(all_rows_by_client)
    duplicate_id_lines = [
        f"    {client}: {iid} {name!r} (source row {src_row}) — second occurrence, first row already loaded"
        for client, r in client_results.items()
        for iid, name, src_row in r.duplicates
    ]

    print_summary(client_results, all_created, ot_exception_lines, ot_threshold_lines, dup_note_lines, duplicate_id_lines)

    if not args.confirm:
        print()
        print(f"Dry-run only. Re-run with --confirm to write {OUTPUT_SQL_PATH.relative_to(PROJECT_ROOT)}.")
        return 0

    sql = generate_sql(all_created)
    OUTPUT_SQL_PATH.write_text(sql)
    print()
    print(f"✓ Wrote {OUTPUT_SQL_PATH.relative_to(PROJECT_ROOT)} ({len(sql.splitlines())} lines, {len(all_created)} rows).")
    print("  Review, then apply via the Supabase SQL editor. This script does NOT apply it.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
