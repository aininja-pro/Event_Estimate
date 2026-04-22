#!/usr/bin/env python3
"""
Sprint 016 — Rate Card Bulk Import.

Reads Dave Morck's cost rate card template and generates SQL that inserts
missing clients, upserts rate card items per (client, fee_type_name), and
seeds a "No Client" rate card from the Audi tab.

Usage:
    python scripts/import_rate_cards.py             # dry-run, prints summary only
    python scripts/import_rate_cards.py --confirm   # also writes scripts/import_rate_cards.sql

Then apply the generated .sql via Supabase SQL editor or psql.

Rules (see planning/sprints/016-rate-card-import/requirements.md):
  1. Skip rows where unit_rate = 0.
  2. Fix typos: 'Detailling' -> 'Detailing', 'Insuarnce' -> 'Insurance'.
  3. Dedupe within a tab by fee_type_name; keep first occurrence.
  4. GL codes come from the fee_types catalog, looked up by name.
  5. Preserve client-level fields (markup_percent, primary_approver) on
     existing clients by never updating the clients row once it exists.

Schema quirk:
  The spreadsheet stores office_cost as a decimal fraction (0.75 = 75%) when
  office_cost_is_percent=TRUE. The app's DB/UI expect whole numbers (75 = 75%).
  The script normalizes by multiplying by 100 on import, with a 0-500 safety
  assertion to catch future data-format changes.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Iterable

import openpyxl

# ---- Paths ----

PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = PROJECT_ROOT / "data" / "imports" / "DriveShop_Cost_Rate_Card_Template.xlsx"
OUTPUT_SQL_PATH = PROJECT_ROOT / "scripts" / "import_rate_cards.sql"

# ---- Constants ----

INSTRUCTIONS_TAB = "Instructions"

TYPO_FIXES = {
    "Detailling": "Detailing",
    "Insuarnce": "Insurance",
}

# Section header -> rate_card_sections.name (verified against supabase_schema.sql)
SECTION_MAP = {
    "Planning & Administration Labor": "Planning & Administration Labor",
    "Onsite Event Labor": "Onsite Event Labor",
    "Creative Costs": "Creative Costs",
    "Logistics Expenses": "Logistics Expenses",
}

EXPECTED_HEADER = (
    "fee_type_name",
    "unit_rate",
    "overtime_rate",
    "corporate_cost",
    "corporate_cost_is_percent",
    "office_cost",
    "office_cost_is_percent",
)

# Per Rule 5, never UPDATE existing client rows. Only INSERT new clients.
# Short-code generator used only when we have to create one.
def default_client_code(name: str) -> str:
    return name.upper().replace(" ", "_")[:32]


# ---- Parsing ----


class ImportRow:
    __slots__ = (
        "fee_type_name",
        "section",
        "unit_rate",
        "overtime_rate",
        "corporate_cost",
        "corporate_cost_is_percent",
        "office_cost",
        "office_cost_is_percent",
        "source_row",
    )

    def __init__(
        self,
        fee_type_name: str,
        section: str,
        unit_rate: float,
        overtime_rate: float | None,
        corporate_cost: float | None,
        corporate_cost_is_percent: bool,
        office_cost: float | None,
        office_cost_is_percent: bool,
        source_row: int,
    ):
        self.fee_type_name = fee_type_name
        self.section = section
        self.unit_rate = unit_rate
        self.overtime_rate = overtime_rate
        self.corporate_cost = corporate_cost
        self.corporate_cost_is_percent = corporate_cost_is_percent
        self.office_cost = office_cost
        self.office_cost_is_percent = office_cost_is_percent
        self.source_row = source_row


class TabResult:
    def __init__(self, tab: str):
        self.tab = tab
        self.rows: list[ImportRow] = []
        self.skipped_zero_rate = 0
        self.deduped: list[str] = []  # names kept-first
        self.typo_fixes: list[tuple[str, str]] = []  # (original, fixed)
        self.warnings: list[str] = []


def _apply_typo_fix(name: str) -> tuple[str, str | None]:
    """Return (fixed_name, typo_key_or_None)."""
    for bad, good in TYPO_FIXES.items():
        if bad in name:
            return name.replace(bad, good), bad
    return name, None


def _normalize_office_cost(raw: float | None, is_percent: bool, tab: str, src_row: int) -> float | None:
    """Scale fractional percent (0.75) to whole-number percent (75). Assert 0-500."""
    if raw is None:
        return None
    if not is_percent:
        return float(raw)
    scaled = float(raw) * 100.0
    if not (0.0 <= scaled <= 500.0):
        raise ValueError(
            f"{tab} r{src_row}: office_cost={raw} scales to {scaled}% which is out of the 0-500 safety range. "
            "This probably means the spreadsheet's office_cost format has changed "
            "(whole numbers instead of fractions). Re-check the normalization rule."
        )
    return scaled


def _is_section_row(row: tuple) -> bool:
    """Section rows have a value in col 1 and nothing else."""
    if not row or row[0] is None:
        return False
    return all(c is None for c in row[1:])


def _as_bool(v) -> bool:
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return v.strip().lower() in ("true", "1", "yes")
    return bool(v)


def _as_float_or_none(v) -> float | None:
    if v is None or v == "":
        return None
    return float(v)


def parse_tab(ws) -> TabResult:
    result = TabResult(ws.title)
    rows_iter = ws.iter_rows(values_only=True)

    header = next(rows_iter, None)
    if not header or tuple(h for h in header[: len(EXPECTED_HEADER)]) != EXPECTED_HEADER:
        result.warnings.append(f"unexpected header: {header!r}")
        return result

    current_section: str | None = None
    seen_names: set[str] = set()
    src_row_idx = 1  # header is row 1

    for row in rows_iter:
        src_row_idx += 1
        # Trim trailing None so section detection works
        while row and row[-1] is None:
            row = row[:-1]
        if not row:
            continue

        if _is_section_row(row):
            raw_section = str(row[0]).strip()
            mapped = SECTION_MAP.get(raw_section)
            if mapped is None:
                result.warnings.append(
                    f"r{src_row_idx}: unknown section {raw_section!r} — items until next section will be skipped"
                )
                current_section = None
            else:
                current_section = mapped
            continue

        if current_section is None:
            result.warnings.append(f"r{src_row_idx}: data row before any section header, skipping")
            continue

        raw_name = str(row[0]).strip()
        fixed_name, typo_key = _apply_typo_fix(raw_name)
        if typo_key:
            result.typo_fixes.append((raw_name, fixed_name))

        try:
            unit_rate = _as_float_or_none(row[1] if len(row) > 1 else None)
            overtime_rate = _as_float_or_none(row[2] if len(row) > 2 else None)
            corporate_cost = _as_float_or_none(row[3] if len(row) > 3 else None)
            corporate_is_pct = _as_bool(row[4]) if len(row) > 4 else False
            office_cost_raw = _as_float_or_none(row[5] if len(row) > 5 else None)
            office_is_pct = _as_bool(row[6]) if len(row) > 6 else False
            office_cost = _normalize_office_cost(office_cost_raw, office_is_pct, ws.title, src_row_idx)
        except (TypeError, ValueError) as e:
            result.warnings.append(f"r{src_row_idx}: could not parse values for {fixed_name!r}: {e}")
            continue

        # Rule 1: skip zero-rate rows
        if unit_rate is None or unit_rate == 0:
            result.skipped_zero_rate += 1
            continue

        # Rule 3: dedupe by fee_type_name, keep first
        if fixed_name in seen_names:
            result.deduped.append(fixed_name)
            continue
        seen_names.add(fixed_name)

        result.rows.append(
            ImportRow(
                fee_type_name=fixed_name,
                section=current_section,
                unit_rate=unit_rate,
                overtime_rate=overtime_rate,
                corporate_cost=corporate_cost,
                corporate_cost_is_percent=corporate_is_pct,
                office_cost=office_cost,
                office_cost_is_percent=office_is_pct,
                source_row=src_row_idx,
            )
        )

    return result


# ---- SQL generation ----


def sql_str(s: str | None) -> str:
    if s is None:
        return "NULL"
    return "'" + s.replace("'", "''") + "'"


def sql_num(v: float | None) -> str:
    return "NULL" if v is None else f"{v:g}"


def sql_bool(v: bool) -> str:
    return "true" if v else "false"


def render_client_upsert(tab: str) -> list[str]:
    """
    Insert client row only if no client (case-insensitive) already exists. Never
    update. Creates with code = UPPERCASE(name). markup_percent & primary_approver
    take DB defaults — AM can set them in the UI later.
    """
    name = tab
    code = default_client_code(tab)
    return [
        f"-- Client: {name}",
        f"INSERT INTO clients (name, code)",
        f"SELECT {sql_str(name)}, {sql_str(code)}",
        f"WHERE NOT EXISTS (SELECT 1 FROM clients WHERE lower(name) = lower({sql_str(name)}));",
        "",
    ]


def render_no_client_seed() -> list[str]:
    return [
        "-- 'No Client' fallback rate card (seeded from Audi tab)",
        "INSERT INTO clients (name, code)",
        "SELECT 'No Client', 'NO_CLIENT'",
        "WHERE NOT EXISTS (SELECT 1 FROM clients WHERE lower(name) = lower('No Client'));",
        "",
    ]


def render_fee_type_insert(name: str, section_key: str, cost_type: str, unit_label: str | None) -> str:
    """
    Insert fee_type if not present. GL code is left NULL — admin sets it in the
    Fee Types admin UI. rate_card_items will link via fee_type_id but carry a
    NULL gl_code until the admin fills it in.
    """
    return (
        f"INSERT INTO fee_types (name, cost_type, section, unit_label) "
        f"SELECT {sql_str(name)}, {sql_str(cost_type)}, {sql_str(section_key)}, {sql_str(unit_label)} "
        f"WHERE NOT EXISTS (SELECT 1 FROM fee_types WHERE name = {sql_str(name)});"
    )


# Dave's four section labels -> (fee_types.section enum, cost_type for fee_types)
SECTION_TO_ENUM = {
    "Planning & Administration Labor": ("planning_admin", "labor"),
    "Onsite Event Labor": ("onsite_labor", "labor"),
    "Creative Costs": ("creative", "labor"),
    "Logistics Expenses": ("logistics", "flat_fee"),
}


def render_item_upsert(client_name: str, item: ImportRow, display_order: int) -> list[str]:
    """
    Upsert rate_card_items row keyed on (client_id, name). UPDATE rate + costs
    on match; INSERT with gl_code/fee_type_id/section_id subqueries on miss.
    Section cost_type drives is_pass_through. Client-level fields untouched.
    """
    name = item.fee_type_name
    section = item.section
    has_ot = item.overtime_rate is not None
    is_pass_through = False  # none of Dave's 4 sections are pass_through

    return [
        "WITH c AS (SELECT id FROM clients WHERE lower(name) = lower(" + sql_str(client_name) + ")),",
        "     s AS (SELECT id FROM rate_card_sections WHERE name = " + sql_str(section) + "),",
        "     ft AS (SELECT id, gl_code, unit_label FROM fee_types WHERE name = " + sql_str(name) + "),",
        "     upd AS (",
        "       UPDATE rate_card_items SET",
        f"         unit_rate = {sql_num(item.unit_rate)},",
        f"         has_overtime_rate = {sql_bool(has_ot)},",
        f"         overtime_rate = {sql_num(item.overtime_rate)},",
        f"         corporate_cost = {sql_num(item.corporate_cost)},",
        f"         corporate_cost_is_percent = {sql_bool(item.corporate_cost_is_percent)},",
        f"         office_cost = {sql_num(item.office_cost)},",
        f"         office_cost_is_percent = {sql_bool(item.office_cost_is_percent)},",
        "         is_active = true",
        "       WHERE client_id = (SELECT id FROM c)",
        "         AND name = " + sql_str(name),
        "       RETURNING id",
        "     )",
        "INSERT INTO rate_card_items (",
        "  client_id, section_id, name, unit_rate, unit_label, gl_code,",
        "  is_from_msa, is_pass_through, has_overtime_rate, overtime_rate,",
        "  overtime_unit_label, overtime_gl_code,",
        "  corporate_cost, corporate_cost_is_percent,",
        "  office_cost, office_cost_is_percent,",
        "  fee_type_id, display_order, is_active, is_rate_locked",
        ")",
        "SELECT",
        "  (SELECT id FROM c), (SELECT id FROM s), " + sql_str(name) + ", "
        + sql_num(item.unit_rate) + ", (SELECT unit_label FROM ft), (SELECT gl_code FROM ft),",
        f"  false, {sql_bool(is_pass_through)}, {sql_bool(has_ot)}, {sql_num(item.overtime_rate)},",
        "  (SELECT unit_label FROM ft), NULL,",
        f"  {sql_num(item.corporate_cost)}, {sql_bool(item.corporate_cost_is_percent)},",
        f"  {sql_num(item.office_cost)}, {sql_bool(item.office_cost_is_percent)},",
        f"  (SELECT id FROM ft), {display_order}, true, false",
        "WHERE NOT EXISTS (SELECT 1 FROM upd);",
        "",
    ]


# ---- Orchestration ----


def parse_workbook() -> tuple[dict[str, TabResult], list[str]]:
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    try:
        results: dict[str, TabResult] = {}
        client_order: list[str] = []
        for tab in wb.sheetnames:
            if tab == INSTRUCTIONS_TAB:
                continue
            results[tab] = parse_tab(wb[tab])
            client_order.append(tab)
        return results, client_order
    finally:
        wb.close()


def collect_unique_fee_types(results: dict[str, TabResult]) -> dict[str, tuple[str, str]]:
    """name -> (section_enum, cost_type). First occurrence wins."""
    catalog: dict[str, tuple[str, str]] = {}
    for tr in results.values():
        for row in tr.rows:
            if row.fee_type_name in catalog:
                continue
            enum_and_cost = SECTION_TO_ENUM.get(row.section)
            if enum_and_cost is not None:
                catalog[row.fee_type_name] = enum_and_cost
    return catalog


def print_summary(results: dict[str, TabResult], client_order: list[str]) -> None:
    total_rows = 0
    total_skipped = 0
    total_deduped = 0
    total_typos = 0
    any_warning = False

    print()
    print("=" * 70)
    print("Sprint 016 — Rate Card Bulk Import — DRY RUN SUMMARY")
    print("=" * 70)
    for tab in client_order:
        tr = results[tab]
        total_rows += len(tr.rows)
        total_skipped += tr.skipped_zero_rate
        total_deduped += len(tr.deduped)
        total_typos += len(tr.typo_fixes)
        print(
            f"  {tab:<14} rows={len(tr.rows):>3}  "
            f"skipped(zero)={tr.skipped_zero_rate:>2}  "
            f"deduped={len(tr.deduped):>1}  "
            f"typos_fixed={len(tr.typo_fixes):>2}"
            + ("  ⚠" if tr.warnings else "")
        )
        for w in tr.warnings:
            any_warning = True
            print(f"      WARN: {w}")

    # Separate "No Client" line (copied from Audi)
    audi = results.get("Audi")
    no_client_rows = len(audi.rows) if audi else 0

    print("-" * 70)
    print(f"  Clients in spreadsheet:   {len(client_order)}")
    print(f"  + No Client (seeded from Audi): {no_client_rows} rows")
    print(f"  Total rate_card_items to upsert: {total_rows + no_client_rows}")
    print(f"  Rows skipped (unit_rate=0): {total_skipped}")
    print(f"  Rows deduped (kept first):  {total_deduped}")
    print(f"  Typo fixes applied:         {total_typos}")

    catalog = collect_unique_fee_types(results)
    print(f"  Unique fee_type names across all tabs: {len(catalog)}")
    print(
        "  (fee_types will be auto-created with NULL gl_code for any names not "
        "already in the catalog — admin sets GL codes in the Fee Types UI.)"
    )
    if any_warning:
        print()
        print("  ⚠ Warnings above — review before running with --confirm.")
    print("=" * 70)


def generate_sql(results: dict[str, TabResult], client_order: list[str]) -> str:
    lines: list[str] = []
    lines.append("-- =============================================================================")
    lines.append("-- Sprint 016 — Rate Card Bulk Import")
    lines.append(f"-- Generated from: {EXCEL_PATH.name}")
    lines.append("-- Safe to re-run: all operations are upserts keyed on name.")
    lines.append("-- =============================================================================")
    lines.append("")
    lines.append("BEGIN;")
    lines.append("")

    # 1. Auto-create any missing fee_types (NULL gl_code for new ones)
    catalog = collect_unique_fee_types(results)
    lines.append("-- Fee types (auto-create only; existing rows untouched)")
    for name in sorted(catalog.keys()):
        section_enum, cost_type = catalog[name]
        lines.append(render_fee_type_insert(name, section_enum, cost_type, None))
    lines.append("")

    # 2. Clients + their rate card items
    for tab in client_order:
        tr = results[tab]
        if not tr.rows:
            continue
        lines.extend(render_client_upsert(tab))
        for i, item in enumerate(tr.rows, start=1):
            lines.extend(render_item_upsert(tab, item, display_order=i))

    # 3. "No Client" rate card (seeded from Audi)
    audi = results.get("Audi")
    if audi and audi.rows:
        lines.extend(render_no_client_seed())
        for i, item in enumerate(audi.rows, start=1):
            lines.extend(render_item_upsert("No Client", item, display_order=i))

    lines.append("COMMIT;")
    lines.append("")
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description="Import Dave Morck's rate card spreadsheet.")
    parser.add_argument(
        "--confirm",
        action="store_true",
        help="Write scripts/import_rate_cards.sql in addition to the dry-run summary.",
    )
    args = parser.parse_args()

    if not EXCEL_PATH.exists():
        print(f"ERROR: Spreadsheet not found at {EXCEL_PATH}", file=sys.stderr)
        return 2

    try:
        results, client_order = parse_workbook()
    except ValueError as e:
        print(f"ERROR during parse: {e}", file=sys.stderr)
        return 1

    print_summary(results, client_order)

    if not args.confirm:
        print()
        print(f"Dry-run only. Re-run with --confirm to write {OUTPUT_SQL_PATH.relative_to(PROJECT_ROOT)}.")
        return 0

    sql = generate_sql(results, client_order)
    OUTPUT_SQL_PATH.write_text(sql)
    print()
    print(f"✓ Wrote {OUTPUT_SQL_PATH.relative_to(PROJECT_ROOT)} ({len(sql.splitlines())} lines).")
    print("  Apply via Supabase SQL editor or `psql $DATABASE_URL < scripts/import_rate_cards.sql`.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
