#!/usr/bin/env python3
"""
Seed fee_types table from the Lucid tab (source of truth for GL codes),
then backfill fee_type_id on rate_card_items.

Also picks up fee types that exist only in non-Lucid clients (from the
existing seed_rate_cards.sql) so every rate_card_item can be linked.

Usage:
    python scripts/seed_fee_types.py > scripts/seed_fee_types.sql

Then run seed_fee_types.sql against Supabase AFTER migration_fee_types.sql.
"""

import openpyxl
import re
import sys

EXCEL_PATH = "data/DriveShop Event Estimate Template_12.01.25.xlsx"
SEED_SQL_PATH = "scripts/seed_rate_cards.sql"

# Section header text -> (section enum, cost_type)
SECTION_MAP = {
    "PLANNING & ADMINISTRATION LABOR": ("planning_admin", "labor"),
    "ONSITE EVENT LABOR": ("onsite_labor", "labor"),
    "TRAVEL EXPENSES (PASS THROUGH)": ("travel", "pass_through"),
    "CREATIVE COSTS": ("creative", "flat_fee"),
    "PRODUCTION EXPENSES (PASS THROUGH)": ("production", "pass_through"),
    "LOGISTICS EXPENSES (FLAT FEE)": ("logistics", "flat_fee"),
}

# rate_card_sections name -> section enum (for parsing seed SQL)
SECTION_NAME_MAP = {
    "Planning & Administration Labor": "planning_admin",
    "Onsite Event Labor": "onsite_labor",
    "Travel Expenses": "travel",
    "Creative Costs": "creative",
    "Production Expenses": "production",
    "Logistics Expenses": "logistics",
}

COST_TYPE_MAP = {
    "planning_admin": "labor",
    "onsite_labor": "labor",
    "travel": "pass_through",
    "creative": "flat_fee",
    "production": "pass_through",
    "logistics": "flat_fee",
}


def escape_sql(s):
    """Escape single quotes for SQL."""
    return s.replace("'", "''")


def normalize_gl_code(val):
    """Convert a raw cell value to a clean GL code string like '4000.01'."""
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    # Handle float precision issues (e.g., 4075.1400000000003 -> 4075.14)
    try:
        f = float(s)
        # GL codes are XXXX.XX format
        return f"{f:.2f}"
    except ValueError:
        return s


def extract_lucid_fee_types():
    """Read the LUCID tab and extract canonical fee types with GL codes."""
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb["LUCID"]

    # Cache all cells
    grid = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                grid[(cell.row, cell.column)] = cell.value

    wb.close()

    fee_types = []  # list of (name, gl_code, section, cost_type, unit_label, display_order)
    current_section = None
    current_cost_type = None
    section_order = 0  # display_order within section

    for r in range(1, ws.max_row + 1):
        col_b = str(grid.get((r, 2), "")).strip()
        col_c = grid.get((r, 3), None)  # unit rate
        col_e = grid.get((r, 5), None)  # GL code in column E

        # Check for section headers
        col_b_upper = col_b.upper()
        matched_section = None
        for header, (sec, ct) in SECTION_MAP.items():
            if col_b_upper.startswith(header):
                matched_section = (sec, ct)
                break

        if matched_section:
            current_section, current_cost_type = matched_section
            section_order = 0
            continue

        # Skip non-item rows
        if not col_b or col_b.startswith("From MSA") or col_b.startswith("Added rates"):
            continue
        if col_b.startswith("Total ") or col_b.startswith("Client Name"):
            continue
        if col_b in ("Third Party Cost Markup", "Agency Fee", "Unit"):
            continue

        # Skip overtime rows (they attach to parent item, not standalone fee types)
        if " OT" in col_b or ">10hrs" in col_b.lower():
            continue

        if current_section is None:
            continue

        section_order += 1
        gl_code = normalize_gl_code(col_e)

        # Determine unit_label from the name or rate context
        unit_label = None
        if "/ hr" in col_b or "/hr" in col_b:
            unit_label = "/ hr"
        elif "/ mnth" in col_b:
            unit_label = "/ mnth"
        elif "/ each" in col_b:
            unit_label = "/ each"
        elif "/ day" in col_b or "/day" in col_b:
            unit_label = "/ day"
        elif "/ event" in col_b or "/event" in col_b:
            unit_label = "/ event"

        fee_types.append((
            col_b,          # name
            gl_code,        # canonical GL code
            current_section,
            current_cost_type,
            unit_label,
            section_order,
        ))

    return fee_types


def extract_non_lucid_fee_types(lucid_names):
    """Find fee type names in seed_rate_cards.sql that don't exist in Lucid."""
    with open(SEED_SQL_PATH) as f:
        content = f.read()

    lines = [l for l in content.split("\n") if l.startswith("INSERT INTO rate_card_items")]

    extra_types = {}  # name -> (section, cost_type, unit_label, gl_code)
    for line in lines:
        sec_match = re.search(r"WHERE name = '([^']+)'\)", line)
        vals_match = re.search(
            r"\), '([^']+)', ([^,]+), ([^,]+), '?([^',]*)'?,\s*(true|false),\s*(true|false)",
            line,
        )
        if not sec_match or not vals_match:
            continue

        section_name = sec_match.group(1)
        name = vals_match.group(1)
        unit_label_raw = vals_match.group(3).strip().strip("'")
        gl_code_raw = vals_match.group(4).strip().strip("'")

        # Skip OT rows
        if " OT" in name or ">10hrs" in name.lower():
            continue

        if name in lucid_names:
            continue

        if name in extra_types:
            continue

        section = SECTION_NAME_MAP.get(section_name)
        if not section:
            continue

        cost_type = COST_TYPE_MAP[section]
        unit_label = unit_label_raw if unit_label_raw and unit_label_raw != "NULL" else None
        gl_code = gl_code_raw if gl_code_raw and gl_code_raw != "NULL" else None
        if gl_code:
            gl_code = normalize_gl_code(gl_code)

        extra_types[name] = (section, cost_type, unit_label, gl_code)

    return extra_types


def generate_sql():
    """Generate the seed SQL file."""
    # Step 1: Extract Lucid fee types (source of truth)
    lucid_fee_types = extract_lucid_fee_types()
    lucid_names = {ft[0] for ft in lucid_fee_types}

    print("-- =============================================================================")
    print("-- Seed data for fee_types master table")
    print("-- Generated from: LUCID tab (source of truth for GL codes)")
    print(f"-- Total Lucid fee types: {len(lucid_fee_types)}")
    print("-- =============================================================================")
    print("-- Run AFTER migration_fee_types.sql.")
    print("-- Safe to re-run: clears fee_type_id references first, then deletes/reinserts.")
    print("-- =============================================================================")
    print()
    print("-- Clear existing references before re-seeding")
    print("UPDATE rate_card_items SET fee_type_id = NULL;")
    print("DELETE FROM fee_types;")
    print()

    # Step 2: Insert Lucid fee types
    print("-- ============================")
    print("-- Fee types from LUCID tab")
    print("-- ============================")

    # Group by section for readability
    current_section = None
    for name, gl_code, section, cost_type, unit_label, display_order in lucid_fee_types:
        if section != current_section:
            current_section = section
            print(f"\n-- Section: {section}")

        gl_sql = f"'{gl_code}'" if gl_code else "NULL"
        unit_sql = f"'{escape_sql(unit_label)}'" if unit_label else "NULL"
        print(
            f"INSERT INTO fee_types (name, gl_code, cost_type, unit_label, section, display_order) "
            f"VALUES ('{escape_sql(name)}', {gl_sql}, '{cost_type}', {unit_sql}, "
            f"'{section}', {display_order});"
        )

    # Step 3: Insert non-Lucid fee types
    extra_types = extract_non_lucid_fee_types(lucid_names)
    if extra_types:
        print()
        print("-- ============================")
        print(f"-- Fee types from other clients (not in Lucid): {len(extra_types)}")
        print("-- GL codes may be NULL or from non-Lucid sources — review these")
        print("-- ============================")

        order_counters = {}
        for name, (section, cost_type, unit_label, gl_code) in sorted(extra_types.items()):
            order_counters[section] = order_counters.get(section, 100) + 1
            gl_sql = f"'{gl_code}'" if gl_code else "NULL"
            unit_sql = f"'{escape_sql(unit_label)}'" if unit_label else "NULL"
            print(
                f"INSERT INTO fee_types (name, gl_code, cost_type, unit_label, section, display_order) "
                f"VALUES ('{escape_sql(name)}', {gl_sql}, '{cost_type}', {unit_sql}, "
                f"'{section}', {order_counters[section]});"
            )

    # Step 4: Backfill fee_type_id on rate_card_items by matching name
    print()
    print("-- ============================")
    print("-- Backfill fee_type_id on rate_card_items")
    print("-- ============================")
    print(
        "UPDATE rate_card_items rci "
        "SET fee_type_id = ft.id "
        "FROM fee_types ft "
        "WHERE rci.name = ft.name;"
    )

    # Step 5: Report unmatched items
    print()
    print("-- Verify: show any rate_card_items that didn't match a fee_type")
    print("-- SELECT rci.name, c.code AS client")
    print("-- FROM rate_card_items rci")
    print("-- JOIN clients c ON c.id = rci.client_id")
    print("-- WHERE rci.fee_type_id IS NULL")
    print("-- ORDER BY c.code, rci.name;")


if __name__ == "__main__":
    generate_sql()
