#!/usr/bin/env python3
"""
Scan every .xlsx file in historical_estimates/ and extract structured data.

Uses read_only=True with iter_rows → grid dict pattern for performance.
NEVER uses ws.cell() in read_only mode (O(n) per access).
"""

import json
import os
import re
import sys
import time
import traceback
from collections import defaultdict

import openpyxl

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUT_DIR = os.path.join(PROJECT_ROOT, "historical_estimates")
OUTPUT = os.path.join(PROJECT_ROOT, "scan_results.json")

TEST_MODE = "--test" in sys.argv
TEST_COUNT = 5

SECTION_HEADERS = {
    "PLANNING & ADMINISTRATION",
    "ACCESS/SPONSORSHIP FEES",
    "VENUE ACCESS & FEES",
    "ONSITE LABOR ACTIVITY",
    "TRAVEL EXPENSES",
    "TRAVEL EXPENSES & FEES",
    "CREATIVE COSTS",
    "CREATIVE EXPENSES",
    "PRODUCTION EXPENSES",
    "MISC EXPENSES",
    "LOGISTICS EXPENSES",
    "OTHER",
    "AGENCY FEES",
}

ALIAS_MAP = {
    "VENUE ACCESS & FEES": "ACCESS/SPONSORSHIP FEES",
    "TRAVEL EXPENSES & FEES": "TRAVEL EXPENSES",
    "CREATIVE EXPENSES": "CREATIVE COSTS",
    "MISC EXPENSES": "PRODUCTION EXPENSES",
}

SKIP_TABS = {"Overview", "Templates", "ROS", "Labor Log", "Template", "Run of Show"}


def load_grid(ws):
    """Load all cell values from a worksheet into a dict of (row, col) -> value."""
    grid = {}
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for col_idx, val in enumerate(row, start=1):
            if val is not None:
                grid[(row_idx, col_idx)] = val
    return grid


def get(grid, row, col):
    """Get value from grid, returns None if not present."""
    return grid.get((row, col))


def safe_float(val):
    """Convert to float if possible."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        cleaned = str(val).replace("$", "").replace(",", "").strip()
        if cleaned:
            return float(cleaned)
    except (ValueError, TypeError):
        pass
    return None


def is_gl_code(val):
    """Check if value looks like a GL code (decimal number like 4000.16)."""
    if val is None:
        return False
    try:
        f = float(val)
        s = str(val)
        return "." in s and 1000 <= f <= 9999.99
    except (ValueError, TypeError):
        return False


def detect_format(wb, sheet_names):
    """Detect FORMAT_A, FORMAT_B, or FORMAT_UNKNOWN."""
    has_overview = "Overview" in sheet_names
    has_templates = any("Templates" in name or "Template" in name for name in sheet_names)

    # Check for GL codes in column A of any non-skip tab
    has_gl_codes = False
    if has_overview and has_templates:
        for name in sheet_names:
            if name in SKIP_TABS or "Template" in name:
                continue
            ws = wb[name]
            grid = load_grid(ws)
            for row in range(1, 200):
                if is_gl_code(get(grid, row, 1)):
                    has_gl_codes = True
                    break
            if has_gl_codes:
                break

    if has_overview and has_templates and has_gl_codes:
        return "FORMAT_A"

    # Check FORMAT_B: P&L summary rows 4-13 with keywords in B or O
    format_b_keywords = {"revenue", "net rev", "per activation total"}
    for name in sheet_names:
        if name in SKIP_TABS:
            continue
        ws = wb[name]
        grid = load_grid(ws)
        for row in range(4, 14):
            for col in (2, 15):  # B=2, O=15
                val = get(grid, row, col)
                if val and str(val).strip().lower() in format_b_keywords:
                    return "FORMAT_B"

    return "FORMAT_UNKNOWN"


def find_sections(grid, max_row):
    """Find section headers in column B, matching exact whitelist only."""
    sections = {}
    for row in range(1, max_row + 1):
        val = get(grid, row, 2)  # Column B
        if val is None:
            continue
        cleaned = str(val).strip().upper()
        if cleaned in SECTION_HEADERS:
            canonical = ALIAS_MAP.get(cleaned, cleaned)
            sections[canonical] = {"start_row": row}
    return sections


def find_section_totals(grid, sections, max_row, fmt):
    """Find total rows and values for each section."""
    result = {}
    for section_name, info in sections.items():
        start_row = info["start_row"]
        total_row = None
        bid_total = None
        recap_total = None

        # Look for "Total ..." row below the section header
        for row in range(start_row + 1, min(start_row + 100, max_row + 1)):
            val = get(grid, row, 2)  # Column B
            if val and str(val).strip().upper().startswith("TOTAL"):
                total_row = row
                bid_total = safe_float(get(grid, row, 8))  # Column H
                if fmt == "FORMAT_A":
                    recap_total = safe_float(get(grid, row, 22))  # Column V
                break
            # If we hit another section header, stop
            if val and str(val).strip().upper() in SECTION_HEADERS:
                break

        result[section_name] = {
            "canonical_name": section_name,
            "section_exists": True,
            "start_row": start_row,
            "total_row": total_row,
            "bid_total": bid_total,
            "recap_total": recap_total,
        }
    return result


def find_grand_total(grid, max_row):
    """Find the Grand Total row and value in column H."""
    for row in range(1, max_row + 1):
        val = get(grid, row, 2)  # Column B
        if val and str(val).strip().upper() == "GRAND TOTAL":
            return safe_float(get(grid, row, 8))  # Column H
    return None


def extract_financials_a(grid):
    """Extract FORMAT_A financial headers from rows 3-7."""
    return {
        "bid_gross": safe_float(get(grid, 4, 11)),      # K4
        "bid_net": safe_float(get(grid, 5, 11)),         # K5
        "bid_margin_dollars": safe_float(get(grid, 4, 12)),  # L4
        "bid_margin_pct": safe_float(get(grid, 4, 13)),  # M4
        "recap_gross": safe_float(get(grid, 4, 24)),     # X4
        "recap_net": safe_float(get(grid, 5, 24)),       # X5
        "recap_margin_dollars": safe_float(get(grid, 4, 25)),  # Y4
        "recap_margin_pct": safe_float(get(grid, 4, 26)),  # Z4
        "payout": safe_float(get(grid, 4, 14)),          # N4
    }


def extract_financials_b(grid):
    """Extract FORMAT_B financial headers from rows 4-9, columns O-P."""
    return {
        "revenue": safe_float(get(grid, 4, 16)),  # P4
        "net_rev": safe_float(get(grid, 5, 16)),   # P5
        "gm": safe_float(get(grid, 7, 16)),        # P7
    }


def extract_labor_roles(grid, sections, fmt):
    """Extract labor roles from ONSITE LABOR ACTIVITY section."""
    roles = []
    labor_section = sections.get("ONSITE LABOR ACTIVITY")
    if not labor_section:
        return roles

    start_row = labor_section["start_row"]
    total_row = labor_section.get("total_row")
    if not total_row:
        # If no total row found, scan up to 80 rows
        end_row = start_row + 80
    else:
        end_row = total_row

    for row in range(start_row + 1, end_row):
        role_name = get(grid, row, 2)  # Column B
        if not role_name:
            continue
        role_str = str(role_name).strip()
        if not role_str or role_str == "---" or role_str.upper() in SECTION_HEADERS:
            continue
        if role_str.upper().startswith("TOTAL"):
            continue

        unit_rate = safe_float(get(grid, row, 5))  # Column E
        if unit_rate is None or unit_rate <= 0:
            continue

        gl_code = None
        if fmt == "FORMAT_A":
            gl_val = get(grid, row, 1)  # Column A
            if gl_val is not None:
                gl_code = str(gl_val).strip()

        cost_rate = safe_float(get(grid, row, 13))  # Column M

        has_ot = bool(re.search(r"OT|>10\s*hrs?", role_str, re.IGNORECASE))

        roles.append({
            "role": role_str,
            "unit_rate": unit_rate,
            "gl_code": gl_code,
            "cost_rate": cost_rate,
            "has_ot_variant": has_ot,
        })

    return roles


def check_recap_data(grid):
    """Check if columns T-Z (20-26) have non-zero numeric values in rows 10-110."""
    for row in range(10, 111):
        for col in range(20, 27):  # T=20 through Z=26
            val = get(grid, row, col)
            if val is not None:
                f = safe_float(val)
                if f is not None and f != 0:
                    return True
    return False


def detect_versions(sheet_names):
    """Look for tabs sharing a common prefix with date suffixes."""
    # Simple heuristic: find tabs that share a long common prefix
    client_tabs = [n for n in sheet_names if n not in SKIP_TABS and "Template" not in n]
    if len(client_tabs) <= 1:
        return []

    # Group by common prefix (at least 5 chars)
    groups = defaultdict(list)
    for i, name1 in enumerate(client_tabs):
        for name2 in client_tabs[i + 1:]:
            # Find longest common prefix
            prefix_len = 0
            for a, b in zip(name1, name2):
                if a == b:
                    prefix_len += 1
                else:
                    break
            if prefix_len >= 5:
                prefix = name1[:prefix_len].rstrip()
                groups[prefix].append(name1)
                groups[prefix].append(name2)

    version_tabs = []
    for prefix, tabs in groups.items():
        unique_tabs = list(set(tabs))
        if len(unique_tabs) >= 2:
            version_tabs.extend(unique_tabs)

    return list(set(version_tabs)) if version_tabs else []


def scan_file(filepath, filename):
    """Scan a single xlsx file and extract all data."""
    result = {"filename": filename, "error": None}

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        result["error"] = str(e)
        return result

    try:
        sheet_names = wb.sheetnames
        result["sheet_names"] = sheet_names

        # Sheet row counts
        sheet_row_counts = {}
        for name in sheet_names:
            ws = wb[name]
            count = 0
            for _ in ws.iter_rows(values_only=True):
                count += 1
            sheet_row_counts[name] = count
        result["sheet_row_counts"] = sheet_row_counts

        # Detect format
        fmt = detect_format(wb, sheet_names)
        result["format"] = fmt

        # Client tabs (not Overview, Templates, ROS, Labor Log, Template, Run of Show)
        client_tabs = [n for n in sheet_names if n not in SKIP_TABS and "Template" not in n]
        result["client_tabs"] = client_tabs

        # Multi-version detection
        version_tabs = detect_versions(sheet_names)
        result["version_tabs"] = version_tabs

        # Find the client tab with the highest grand total (actual data)
        main_grid = None
        main_max_row = 0
        main_tab = None
        sections = {}
        best_grand_total = 0

        for tab_name in client_tabs:
            ws = wb[tab_name]
            grid = load_grid(ws)
            max_row = max((r for r, c in grid.keys()), default=0) if grid else 0
            raw_sections = find_sections(grid, max_row)
            if not raw_sections:
                continue
            gt = find_grand_total(grid, max_row)
            gt_val = gt if gt and gt > 0 else 0
            if gt_val > best_grand_total or main_grid is None:
                best_grand_total = gt_val
                main_grid = grid
                main_max_row = max_row
                main_tab = tab_name
                sections = find_section_totals(grid, raw_sections, max_row, fmt)

        result["main_tab"] = main_tab

        if main_grid is not None:
            result["sections"] = sections
            result["grand_total"] = find_grand_total(main_grid, main_max_row)

            if fmt == "FORMAT_A":
                result["financials"] = extract_financials_a(main_grid)
            elif fmt == "FORMAT_B":
                result["financials"] = extract_financials_b(main_grid)
            else:
                result["financials"] = {}

            result["labor_roles"] = extract_labor_roles(main_grid, sections, fmt)
            result["has_recap_data"] = check_recap_data(main_grid)
        elif client_tabs:
            # No sections found, use first client tab for financials anyway
            ws = wb[client_tabs[0]]
            grid = load_grid(ws)
            max_row = max((r for r, c in grid.keys()), default=0) if grid else 0
            result["sections"] = {}
            result["grand_total"] = find_grand_total(grid, max_row)

            if fmt == "FORMAT_A":
                result["financials"] = extract_financials_a(grid)
            elif fmt == "FORMAT_B":
                result["financials"] = extract_financials_b(grid)
            else:
                result["financials"] = {}

            result["labor_roles"] = []
            result["has_recap_data"] = check_recap_data(grid)
        else:
            result["sections"] = {}
            result["grand_total"] = None
            result["financials"] = {}
            result["labor_roles"] = []
            result["has_recap_data"] = False

    except Exception as e:
        result["error"] = str(e)
        result["error_traceback"] = traceback.format_exc()
    finally:
        wb.close()

    return result


def main():
    files = sorted([f for f in os.listdir(INPUT_DIR) if f.endswith(".xlsx")])
    print(f"Found {len(files)} xlsx files to scan")

    if TEST_MODE:
        files = files[:TEST_COUNT]
        print(f"TEST MODE: scanning first {TEST_COUNT} files")

    results = []
    errors = []
    format_counts = defaultdict(int)
    start_time = time.time()

    for i, filename in enumerate(files, 1):
        filepath = os.path.join(INPUT_DIR, filename)
        result = scan_file(filepath, filename)
        results.append(result)

        if result.get("error"):
            errors.append((filename, result["error"]))
        else:
            format_counts[result.get("format", "UNKNOWN")] += 1

        if i % 50 == 0 or TEST_MODE:
            elapsed = time.time() - start_time
            rate = i / elapsed if elapsed > 0 else 0
            print(f"[{i}/{len(files)}] {rate:.1f} files/sec — {filename[:60]}")

    elapsed = time.time() - start_time

    output = {
        "total_scanned": len(results),
        "errors": len(errors),
        "format_counts": dict(format_counts),
        "elapsed_seconds": round(elapsed, 1),
        "results": results,
    }

    with open(OUTPUT, "w") as f:
        json.dump(output, f, indent=2, default=str)

    print(f"\n--- RESULTS ---")
    print(f"Scanned: {len(results)} files in {elapsed:.1f}s")
    print(f"Format counts: {dict(format_counts)}")
    print(f"Errors: {len(errors)}")
    if errors:
        print("Error files:")
        for fn, err in errors[:10]:
            print(f"  {fn}: {err[:100]}")
    print(f"Output: {OUTPUT}")

    # Quick stats
    with_sections = sum(1 for r in results if r.get("sections"))
    with_labor = sum(1 for r in results if r.get("labor_roles"))
    with_recap = sum(1 for r in results if r.get("has_recap_data"))
    total_roles = sum(len(r.get("labor_roles", [])) for r in results)
    print(f"\nFiles with sections: {with_sections}")
    print(f"Files with labor roles: {with_labor}")
    print(f"Files with recap data: {with_recap}")
    print(f"Total labor role rows across all files: {total_roles}")


if __name__ == "__main__":
    main()
